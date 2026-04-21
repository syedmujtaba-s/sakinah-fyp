"""
Admin endpoints for managing Seerah stories.

- POST /api/admin/stories       — Add a single story via JSON
- POST /api/admin/upload         — Bulk upload stories from Excel (.xlsx)
- GET  /api/admin/stories        — List all stories
- DELETE /api/admin/stories/{id} — Delete a story by ID
"""

import io
import os
from typing import Optional
from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import FileResponse
from pydantic import BaseModel, field_validator
from bson import ObjectId
from openpyxl import Workbook, load_workbook

from database import stories_collection, story_index_collection
from rag import sync_new_stories

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "story_template.xlsx")

router = APIRouter(prefix="/api/admin", tags=["admin"])

# --- Supported emotions (must match guidance_router) ---
SUPPORTED_EMOTIONS = [
    "happy", "sad", "anxious", "angry", "confused",
    "grateful", "lonely", "stressed", "fearful", "guilty",
    "hopeless", "overwhelmed", "rejected", "embarrassed", "lost"
]


# ============================
#  Pydantic Models
# ============================
class StoryCreate(BaseModel):
    title: str
    period: str
    emotions: list[str]
    summary: str
    story: str
    lessons: list[str]
    practical_advice: list[str]

    @field_validator("emotions")
    @classmethod
    def validate_emotions(cls, v):
        for e in v:
            if e.lower() not in SUPPORTED_EMOTIONS:
                raise ValueError(f"Unsupported emotion: '{e}'. Supported: {SUPPORTED_EMOTIONS}")
        return [e.lower() for e in v]

    @field_validator("lessons", "practical_advice")
    @classmethod
    def at_least_one(cls, v, info):
        if len(v) < 1:
            raise ValueError(f"{info.field_name} must have at least 1 item")
        return v


# ============================
#  1. ADD SINGLE STORY (JSON)
# ============================
@router.post("/stories")
async def add_story(story: StoryCreate):
    """Add a single story via JSON body. Auto-indexes it for RAG."""

    # Check for duplicate title
    existing = stories_collection.find_one({"title": story.title})
    if existing:
        raise HTTPException(status_code=409, detail=f"Story with title '{story.title}' already exists.")

    doc = story.model_dump()
    result = stories_collection.insert_one(doc)

    # Auto-index the new story for RAG search
    sync_new_stories()

    return {
        "message": "Story added and indexed successfully",
        "story_id": str(result.inserted_id),
        "title": story.title
    }


# ============================
#  2. BULK UPLOAD FROM EXCEL
# ============================
@router.post("/upload")
async def upload_stories(file: UploadFile = File(...)):
    """
    Upload an Excel (.xlsx) file to bulk-add stories.

    Expected columns:
    title | period | emotions | summary | story | lesson_1 | lesson_2 | lesson_3 | advice_1 | advice_2 | advice_3

    - emotions column: comma-separated (e.g. "sad, lonely, hopeless")
    - Empty lesson/advice cells are skipped
    """

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted.")

    contents = await file.read()

    try:
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    # Read header row
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty.")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    # Validate required columns exist
    required = {"title", "period", "emotions", "summary", "story"}
    found = set(headers)
    missing = required - found
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {missing}. Found: {headers}"
        )

    added = []
    skipped = []
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        row_dict = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            row_dict[header] = str(value).strip() if value else ""

        title = row_dict.get("title", "")
        if not title:
            skipped.append({"row": row_num, "reason": "Empty title"})
            continue

        # Check duplicate
        if stories_collection.find_one({"title": title}):
            skipped.append({"row": row_num, "reason": f"Duplicate title: '{title}'"})
            continue

        # Parse emotions (comma-separated)
        raw_emotions = row_dict.get("emotions", "")
        emotions = [e.strip().lower() for e in raw_emotions.split(",") if e.strip()]

        # Validate emotions
        invalid_emotions = [e for e in emotions if e not in SUPPORTED_EMOTIONS]
        if invalid_emotions:
            errors.append({"row": row_num, "title": title, "error": f"Invalid emotions: {invalid_emotions}"})
            continue

        if not emotions:
            errors.append({"row": row_num, "title": title, "error": "No emotions provided"})
            continue

        # Collect lessons (lesson_1, lesson_2, lesson_3, ...)
        lessons = []
        for key in sorted(headers):
            if key.startswith("lesson") and row_dict.get(key):
                lessons.append(row_dict[key])

        # Collect advice (advice_1, advice_2, advice_3, ...)
        practical_advice = []
        for key in sorted(headers):
            if key.startswith("advice") and row_dict.get(key):
                practical_advice.append(row_dict[key])

        if not lessons:
            errors.append({"row": row_num, "title": title, "error": "At least 1 lesson required"})
            continue

        if not practical_advice:
            errors.append({"row": row_num, "title": title, "error": "At least 1 advice required"})
            continue

        doc = {
            "title": title,
            "period": row_dict.get("period", ""),
            "emotions": emotions,
            "summary": row_dict.get("summary", ""),
            "story": row_dict.get("story", ""),
            "lessons": lessons,
            "practical_advice": practical_advice,
        }

        stories_collection.insert_one(doc)
        added.append({"row": row_num, "title": title})

    # Auto-index all new stories for RAG
    if added:
        sync_new_stories()

    wb.close()

    return {
        "message": f"Upload complete. {len(added)} added, {len(skipped)} skipped, {len(errors)} errors.",
        "added": added,
        "skipped": skipped,
        "errors": errors,
    }


# ============================
#  3. LIST ALL STORIES
# ============================
@router.get("/stories")
async def list_stories():
    """List all stories in the database."""
    stories = []
    for s in stories_collection.find({}).sort("title", 1):
        s["_id"] = str(s["_id"])
        stories.append(s)
    return {"count": len(stories), "stories": stories}


@router.get("/stories/count")
async def stories_count():
    """Return only the total story count — cheap for dashboards."""
    return {"count": stories_collection.count_documents({})}


# ============================
#  4. UPDATE A STORY
# ============================
@router.put("/stories/{story_id}")
async def update_story(story_id: str, story: StoryCreate):
    """Update a story and re-index it for RAG."""
    try:
        obj_id = ObjectId(story_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid story ID format.")

    existing = stories_collection.find_one({"_id": obj_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Story not found.")

    # Title uniqueness — allow keeping the same title, block collision with a different story
    dup = stories_collection.find_one({"title": story.title, "_id": {"$ne": obj_id}})
    if dup:
        raise HTTPException(
            status_code=409,
            detail=f"Another story with title '{story.title}' already exists.",
        )

    stories_collection.update_one({"_id": obj_id}, {"$set": story.model_dump()})

    # Drop stale index entries and rebuild chunks for this story
    story_index_collection.delete_many({"story_id": story_id})
    sync_new_stories()

    return {
        "message": "Story updated and re-indexed successfully",
        "story_id": story_id,
        "title": story.title,
    }


# ============================
#  5. DELETE A STORY
# ============================
@router.delete("/stories/{story_id}")
async def delete_story(story_id: str):
    """Delete a story and its index entry."""
    try:
        result = stories_collection.delete_one({"_id": ObjectId(story_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid story ID format.")

    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Story not found.")

    # Remove from index too
    story_index_collection.delete_many({"story_id": story_id})

    return {"message": "Story and its index entry deleted.", "story_id": story_id}


# ============================
#  6. DOWNLOAD EXCEL TEMPLATE
# ============================
@router.get("/template")
async def template_info():
    """Returns the expected column format for the Excel upload (JSON)."""
    return {
        "message": "Use these exact column headers in your Excel file (row 1):",
        "columns": [
            "title", "period", "emotions", "summary", "story",
            "lesson_1", "lesson_2", "lesson_3",
            "advice_1", "advice_2", "advice_3"
        ],
        "notes": {
            "emotions": "Comma-separated, e.g.: sad, lonely, hopeless",
            "lesson/advice": "Leave cells empty if fewer than 3. At least 1 of each required.",
            "supported_emotions": SUPPORTED_EMOTIONS
        }
    }


@router.get("/template/download")
async def download_template_file():
    """Returns the pre-built story_template.xlsx file for admins to fill in."""
    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(status_code=404, detail="Template file not found on server.")
    return FileResponse(
        TEMPLATE_PATH,
        filename="story_template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
